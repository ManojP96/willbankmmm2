import streamlit as st
from numerize.numerize import numerize
import io
import pandas as pd
from utilities import (format_numbers,decimal_formater,
                       channel_name_formating,
                       load_local_css,set_header,
                       initialize_data,
                       load_authenticator)
from openpyxl import Workbook
from openpyxl.styles import Alignment,Font,PatternFill
import pickle
import streamlit_authenticator as stauth
import yaml
from yaml import SafeLoader
from classes import class_from_dict

st.set_page_config(layout='wide')
load_local_css('styles.css')
set_header()

# for k, v in st.session_state.items():
#     if k not in ['logout', 'login','config'] and not k.startswith('FormSubmitter'):
#         st.session_state[k] = v

def create_scenario_summary(scenario_dict):
    summary_rows = []
    for channel_dict in scenario_dict['channels']:
        name_mod = channel_name_formating(channel_dict['name'])
        summary_rows.append([name_mod,
                             channel_dict.get('actual_total_spends') * channel_dict.get('conversion_rate'),
                             channel_dict.get('modified_total_spends') * channel_dict.get('conversion_rate'),
                             channel_dict.get('actual_total_sales') ,
                             channel_dict.get('modified_total_sales'),
                             channel_dict.get('actual_total_sales') / (channel_dict.get('actual_total_spends') * channel_dict.get('conversion_rate')), 
                             channel_dict.get('modified_total_sales') / (channel_dict.get('modified_total_spends') * channel_dict.get('conversion_rate')),
                             channel_dict.get('actual_mroi'), 
                             channel_dict.get('modified_mroi'),
                             channel_dict.get('actual_total_spends') * channel_dict.get('conversion_rate') / channel_dict.get('actual_total_sales'),
                             channel_dict.get('modified_total_spends') * channel_dict.get('conversion_rate') / channel_dict.get('modified_total_sales')])
        
    summary_rows.append(['Total',
                         scenario_dict.get('actual_total_spends'), 
                         scenario_dict.get('modified_total_spends'),
                         scenario_dict.get('actual_total_sales'), 
                         scenario_dict.get('modified_total_sales'),
                         scenario_dict.get('actual_total_sales') / scenario_dict.get('actual_total_spends'),
                         scenario_dict.get('modified_total_sales') / scenario_dict.get('modified_total_spends'),
                         '-',
                         '-',
                         scenario_dict.get('actual_total_spends') / scenario_dict.get('actual_total_sales'),
                         scenario_dict.get('modified_total_spends') / scenario_dict.get('modified_total_sales')])
    
    columns_index = pd.MultiIndex.from_product([[''],['Channel']], names=["first", "second"])
    columns_index = columns_index.append(pd.MultiIndex.from_product([['Spends','NRPU','ROI','MROI','Spend per NRPU'],['Actual','Simulated']], names=["first", "second"]))
    return  pd.DataFrame(summary_rows, columns=columns_index)
    

   
def summary_df_to_worksheet(df, ws):
    heading_fill = PatternFill(fill_type='solid',start_color='15C39A',end_color='15C39A')
    for j,header in enumerate(df.columns.values):
        col = j + 1
        for i in range(1,3):
            ws.cell(row=i,column=j+1,value=header[i-1]).font = Font(bold=True, color='#11B6BD')
            ws.cell(row=i,column=j+1).fill = heading_fill
        if  col > 1 and (col - 6)%5==0:    
            ws.merge_cells(start_row=1, end_row=1, start_column = col-3, end_column=col)
            ws.cell(row=1,column=col).alignment = Alignment(horizontal='center')
    for i,row in enumerate(df.itertuples()):
        for j,value in enumerate(row):
            if j == 0:
                continue
            elif (j-2)%4 == 0 or (j-3)%4 == 0:
                ws.cell(row=i+3, column = j, value=value).number_format = '$#,##0.0' 
            else:
                ws.cell(row=i+3, column = j, value=value)
   
def scenario_df_to_worksheet(df, ws):
    heading_fill = PatternFill(fill_type='solid',start_color='#11B6BD',end_color='#11B6BD')
    for j,header in enumerate(df.columns.values):
        ws.cell(row=1,column=j+1,value=header).font = Font(bold=True, color='#11B6BD')
        ws.cell(row=1,column=j+1).fill = heading_fill
    for i,row in enumerate(df.itertuples()):
        for j,value in enumerate(row):
            if j == 0:
                continue
            elif j == 1:
                ws.cell(row=i+2, column = j, value=value)
            else:
                ws.cell(row=i+2, column = j, value=value).number_format = '$#,##0.0' 
   
def download_scenarios():
    """
    Makes a excel with all saved scenarios and saves it locally
    """
    ## create summary page
    if len(scenarios_to_download) == 0:
        return
    wb = Workbook()
    wb.iso_dates = True
    wb.remove(wb.active)
    st.session_state['xlsx_buffer'] = io.BytesIO()
    summary_df = None
    print(scenarios_to_download)
    for scenario_name in scenarios_to_download:
        scenario_dict =  st.session_state['saved_scenarios'][scenario_name]
        _spends = []
        column_names = ['Date']
        _sales = None
        dates = None
        summary_rows = []
        for channel in scenario_dict['channels']:
            if dates is None:
                dates = channel.get('dates')
                _spends.append(dates)
            if _sales is None:
                _sales = channel.get('modified_sales')
            else:
                _sales += channel.get('modified_sales')
            _spends.append(channel.get('modified_spends') * channel.get('conversion_rate'))
            column_names.append(channel.get('name'))
            
            name_mod = channel_name_formating(channel['name'])
            summary_rows.append([name_mod,
                                channel.get('modified_total_spends') * channel.get('conversion_rate') ,
                                channel.get('modified_total_sales'),
                                channel.get('modified_total_sales') / channel.get('modified_total_spends') * channel.get('conversion_rate'),
                                channel.get('modified_mroi'),
                                channel.get('modified_total_sales') / channel.get('modified_total_spends') * channel.get('conversion_rate')])
        _spends.append(_sales)
        column_names.append('NRPU')
        scenario_df = pd.DataFrame(_spends).T
        scenario_df.columns = column_names
        ## write to sheet
        ws = wb.create_sheet(scenario_name)
        scenario_df_to_worksheet(scenario_df, ws)    
        summary_rows.append(['Total',
                        scenario_dict.get('modified_total_spends') ,
                        scenario_dict.get('modified_total_sales'),
                        scenario_dict.get('modified_total_sales') / scenario_dict.get('modified_total_spends'),
                        '-',
                        scenario_dict.get('modified_total_spends') / scenario_dict.get('modified_total_sales')])
        columns_index = pd.MultiIndex.from_product([[''],['Channel']], names=["first", "second"])
        columns_index = columns_index.append(pd.MultiIndex.from_product([[scenario_name],['Spends','NRPU','ROI','MROI','Spends per NRPU']], names=["first", "second"]))
        if summary_df is None:
            summary_df = pd.DataFrame(summary_rows, columns = columns_index)
            summary_df = summary_df.set_index(('','Channel'))
        else:
            _df = pd.DataFrame(summary_rows, columns = columns_index)
            _df = _df.set_index(('','Channel'))
            summary_df = summary_df.merge(_df, left_index=True, right_index=True)
    ws = wb.create_sheet('Summary',0)
    summary_df_to_worksheet(summary_df.reset_index(), ws)
    wb.save(st.session_state['xlsx_buffer'])
    st.session_state['disable_download_button'] = False

def disable_download_button():
    st.session_state['disable_download_button'] =True

def transform(x):
    if x.name == ("",'Channel'):
        return x
    elif x.name[0] == 'ROI' or x.name[0] == 'MROI':
        return x.apply(lambda y : y if isinstance(y,str) else decimal_formater(format_numbers(y,include_indicator=False,n_decimals=4),n_decimals=4))
    else:
        return x.apply(lambda y : y if isinstance(y,str) else format_numbers(y))

def delete_scenario():
    if selected_scenario in st.session_state['saved_scenarios']:
        del st.session_state['saved_scenarios'][selected_scenario]
        with open('../saved_scenarios.pkl', 'wb') as f:
            pickle.dump(st.session_state['saved_scenarios'],f)     
            
def load_scenario():
    if selected_scenario in st.session_state['saved_scenarios']:
        st.session_state['scenario'] = class_from_dict(selected_scenario_details)
        


authenticator = st.session_state.get('authenticator')
if authenticator is None:
    authenticator = load_authenticator()

name, authentication_status, username = authenticator.login('Login', 'main')
auth_status = st.session_state.get('authentication_status')

if auth_status == True:
    is_state_initiaized = st.session_state.get('initialized',False)
    if not is_state_initiaized:
        print("Scenario page state reloaded")
        initialize_data()


    saved_scenarios = st.session_state['saved_scenarios']


    if len(saved_scenarios) ==0:
        st.header('No saved scenarios')
        
    else:
        
        with st.sidebar:
            selected_scenario = st.radio(
                'Pick a scenario to view details',
                list(saved_scenarios.keys())
            )
            st.markdown("""<hr>""", unsafe_allow_html=True)
            scenarios_to_download = st.multiselect('Select scenarios to download',
                        list(saved_scenarios.keys()))
            
            st.button('Prepare download',on_click=download_scenarios)
            st.download_button(
                    label="Download Scenarios",
                    data=st.session_state['xlsx_buffer'].getvalue(),
                    file_name="scenarios.xlsx",
                    mime="application/vnd.ms-excel",
                    disabled= st.session_state['disable_download_button'],
                    on_click= disable_download_button
                )
            
        column_1, column_2,column_3 = st.columns((6,1,1))
        with column_1:
            st.header(selected_scenario)
        with column_2:
            st.button('Delete scenarios', on_click=delete_scenario)
        with column_3:
            st.button('Load Scenario', on_click=load_scenario)
        
        selected_scenario_details = saved_scenarios[selected_scenario]
        
        pd.set_option('display.max_colwidth', 100)
            
        st.markdown(create_scenario_summary(selected_scenario_details).transform(transform).style.set_table_styles(
    [{
        'selector': 'th',
        'props': [('background-color', '#11B6BD')]
    },
        {
        'selector' : 'tr:nth-child(even)',
        'props' : [('background-color', '#11B6BD')]
        }
        ]).to_html(),unsafe_allow_html=True)
        
elif auth_status == False:
    st.error('Username/Password is incorrect')
    
if auth_status != True:
    try:
        username_forgot_pw, email_forgot_password, random_password = authenticator.forgot_password('Forgot password')
        if username_forgot_pw:
            st.success('New password sent securely')
            # Random password to be transferred to user securely
        elif username_forgot_pw == False:
            st.error('Username not found')
    except Exception as e:
        st.error(e)
